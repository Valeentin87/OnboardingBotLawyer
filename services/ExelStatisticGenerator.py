from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from datetime import datetime
from typing import Dict, List, Tuple, Any
import os
from pathlib import Path

from bot.adapters.max.create_bot import logger

class ExcelStatisticGenerator:
    """
    Класс для генерации Excel-документа с данными о прохождении обучения пользователями.
    """

    def __init__(self, statistic_data: Dict[Tuple[str, str], Dict[str, Any]]):
        """
        Инициализация класса с данными статистики.

        :param statistic_data: словарь с данными о прохождении обучения, где ключ — кортеж (ID, Имя Фамилия)
        """
        self.statistic_data = statistic_data

    def _format_date(self, date_str: str) -> str:
        """
        Форматирует дату из формата ISO в формат ДД.ММ.ГГГГ.
        :param date_str: строка с датой в формате ISO (например, '2026-04-20T06:26:13.151963')
        :return: отформатированная строка даты (например, '20.04.2026')
        """
        try:
            dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
            return dt.strftime('%d.%m.%Y')
        except (ValueError, TypeError):
            return "Попытка не завершена"

    # def _extract_data(self) -> List[List[Any]]:
    #     """
    #     Извлекает и структурирует данные из словаря статистики.
    #     :return: список строк для записи в Excel (каждая строка — список значений)
    #     """
    #     rows = []

    #     for (user_id, full_name), user_data in self.statistic_data.items():
    #         # Обработка завершённых попыток
    #         for course_data in user_data['completed_attemps']:
    #             for course_name, data in course_data.items():
    #                 lessons_completed = data['lessons_completed']
    #                 total_lessons = lessons_completed  # Для завершённых курсов общее число уроков = пройденным
    #                 accuracy = data['accuracy_percent']
    #                 date_completed = self._format_date(data['date_completed'])

    #                 rows.append([
    #             full_name,
    #             user_id,
    #             course_name,
    #             f"{lessons_completed}/{total_lessons}",
    #             f"{accuracy}%",
    #             date_completed
    #         ])

    #         # Обработка незавершённых попыток
    #         for course_data in user_data['not_completed_courses']:
    #             for course_name, data in course_data.items():
    #                 lessons_completed = data['lessons_completed']
    #                 total_lessons = data['total_lessons']
    #                 accuracy = data['accuracy_percent']

    #                 rows.append([
    #                     full_name,
    #                     user_id,
    #                     course_name,
    #                     f"{lessons_completed}/{total_lessons}",
    #                     f"{accuracy}%",
    #                     "Попытка не завершена"
    #                 ])

    #     # Сортировка: сначала по имени (алфавитный порядок), затем по дате (с учётом "Попытка не завершена")
    #     def sort_key(row):
    #         name = row[0]
    #         date = row[5]
    #         # Помещаем строки с "Попытка не завершена" в конец для каждого пользователя
    #         date_sort = float('inf') if date == "Попытка не завершена" else datetime.strptime(date, '%d.%m.%Y') if date != "Попытка не завершена" else None
    #         return (name, date_sort)

    #     rows.sort(key=sort_key)
    #     return rows

    
    def _extract_data(self) -> List[List[Any]]:
        """
        Извлекает и структурирует данные из словаря статистики.
        :return: список строк для записи в Excel (каждая строка — список значений)
        """
        rows = []

        for (user_id, full_name), user_data in self.statistic_data.items():
            # Обработка завершённых попыток
            for course_data in user_data['completed_attemps']:
                for course_name, data in course_data.items():
                    print(f'На этапе формирования итоговых данных для exel документа {course_name=}')
                    lessons_completed = data['lessons_completed']
                    total_lessons = lessons_completed  # Для завершённых курсов общее число уроков = пройденным
                    accuracy = data['accuracy_percent']
                    date_completed = self._format_date(data['date_completed'])
                    
                    if course_name == 'Другой сотрудник':
                        course_name = 'Обучение по продукту'

                    rows.append([
                        full_name,
                        user_id,
                        course_name,
                        f"{lessons_completed if course_name == 'Обучение по продажам' else 7}/{total_lessons if course_name == 'Обучение по продажам' else 7}",
                        accuracy,
                date_completed
            ])

            # Обработка незавершённых попыток
            for course_data in user_data['not_completed_courses']:
                for course_name, data in course_data.items():
                    print(f'На этапе формирования итоговых данных для exel документа {course_name=}')
                    lessons_completed = data['lessons_completed']
                    total_lessons = data['total_lessons']
                    accuracy = data['accuracy_percent']

                    if course_name == 'Другой сотрудник':
                        course_name = 'Обучение по продукту'
                    
                    rows.append([
                        full_name,
                        user_id,
                        course_name,
                        f"{lessons_completed}/{total_lessons if course_name == 'Обучение по продажам' else 7}",
                        accuracy,
                        "Попытка не завершена"
                    ])

        # Сортировка: сначала по имени (алфавитный порядок), затем по дате
        def sort_key(row):
            name = row[0]
            date_str = row[5]

            if date_str == "Попытка не завершена":
                # Для незавершённых попыток используем большое число (конец сортировки)
                date_timestamp = float('inf')
            else:
                # Преобразуем строку даты в timestamp для корректного сравнения
                try:
                    date_obj = datetime.strptime(date_str, '%d.%m.%Y')
                    date_timestamp = date_obj.timestamp()
                except ValueError:
                    # Если формат даты некорректен, помещаем в конец
                    date_timestamp = float('inf')

            return (name, date_timestamp)

        rows.sort(key=sort_key)
        return rows

    
    
    def generate_excel(self, filename: str) -> None:
        """
        Создаёт Excel-файл с данными статистики обучения.
        :param filename: имя файла для сохранения (с расширением .xlsx)
        """
        logger.info("Стартовал")

        # Получаем абсолютный путь
        abs_filename = Path(filename).resolve()
        directory = abs_filename.parent
        
        try:
            # Создаём директорию, если её нет, с правами на запись
            directory.mkdir(parents=True, exist_ok=True)
            logger.info(f"Директория обеспечена: {directory}")

            # Проверяем права записи
            if not os.access(directory, os.W_OK):
                # Пытаемся изменить права (Linux/macOS)
                try:
                    directory.chmod(0o755)
                    logger.info(f"Права директории изменены на 755: {directory}")
                except PermissionError:
                    raise PermissionError(
                        f"Нет прав на запись и изменение прав в директории: {directory}"
                    )        
        
            # Создаём новую книгу и активный лист
            wb = Workbook()
            ws = wb.active
            ws.title = "Статистика обучения"
            
            center_alignment = Alignment(horizontal='center', vertical='center')

            # Заголовки таблицы (полужирный шрифт)
            headers = [
                "Имя Фамилия",
                "ID",
                "Курс",
                "Уроков",
                "Точность",
                "Дата"
            ]

            # Добавляем заголовки в первую строку
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = center_alignment

            # Извлекаем данные и заполняем таблицу
            data_rows = self._extract_data()
            for row_idx, row_data in enumerate(data_rows, start=2):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # Автоподбор ширины столбцов
            for num_column, column in enumerate(ws.columns):
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if num_column in [1, 3, 4]:
                        cell.alignment = center_alignment
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем файл
            logger.info("Перед сохранением")
            wb.save(filename)
            logger.info(f"Excel-файл успешно создан: {filename}")
        
        except PermissionError as e:
            logger.error(f"Ошибка доступа: {e}")
            raise
        except Exception as e:
            logger.error(f"Ошибка при создании Excel: {e}")
            raise


if __name__ == "__main__":
    

    test_data = {('108241884', 'Мельникова Анна'): {'completed_attemps': [],
                                    'not_completed_courses': [{'Обучение по продажам': {'accuracy_percent': 100.0,
                                                                                        'correct_answers': 10,
                                                                                        'lessons_completed': 2,
                                                                                        'total_answers': 10,
                                                                                        'total_lessons': 43}}]},
 ('121356085', 'Галиуллин Айдар'): {'completed_attemps': [],
                                    'not_completed_courses': [{'Обучение по продажам': {'accuracy_percent': 80.0,
                                                                                        'correct_answers': 8,
                                                                                        'lessons_completed': 2,
                                                                                        'total_answers': 10,
                                                                                        'total_lessons': 43}}]},
 ('152163122', 'Галикаев Артур'): {'completed_attemps': [],
                                   'not_completed_courses': [{'Обучение по продажам': {'accuracy_percent': 93.3,
                                                                                       'correct_answers': 42,
                                                                                       'lessons_completed': 7,
                                                                                       'total_answers': 45,
                                                                                       'total_lessons': 43}}]},
 ('175082514', 'Пользователь Тестовый'): {'completed_attemps': [],
                                          'not_completed_courses': [{'Обучение по продажам': {'accuracy_percent': 26.7,
                                                                                              'correct_answers': 12,
                                                                                              'lessons_completed': 7,
                                                                                              'total_answers': 45,
                                                                                              'total_lessons': 43}}]},
 ('20759321', 'Гаранин Алексей'): {'completed_attemps': [{'Обучение по продажам': {'accuracy_percent': 76.2,
                                                                                   'correct_answers': 179,
                                                                                   'course_name': 'Обучение '
                                                                                                  'по '
                                                                                                  'продажам',
                                                                                   'date_completed': '2026-04-20T06:26:13.151963',
                                                                                   'lessons_completed': 43,
                                                                                   'total_answers': 235}}],
                                   'not_completed_courses': []},
 ('209642009', 'Самарханова Карина'): {'completed_attemps': [],
                                       'not_completed_courses': [{'Обучение по продажам': {'accuracy_percent': 100.0,
                                                                                           'correct_answers': 5,
                                                                                           'lessons_completed': 1,
                                                                                           'total_answers': 5,
                                                                                           'total_lessons': 43}}]},
 ('228312484', 'Соколова Ирина'): {'completed_attemps': [],
                                   'not_completed_courses': [{'Обучение по продажам': {'accuracy_percent': 91.1,
                                                                                       'correct_answers': 41,
                                                                                       'lessons_completed': 7,
                                                                                       'total_answers': 45,
                                                                                       'total_lessons': 43}}]},
 ('24297191', 'Малыхин Сергей'): {'completed_attemps': [],
                                  'not_completed_courses': [{'Обучение по продажам': {'accuracy_percent': 93.3,
                                                                                      'correct_answers': 28,
                                                                                      'lessons_completed': 6,
                                                                                      'total_answers': 30,
                                                                                      'total_lessons': 43}}]},
 ('276950556', 'Табакова Татьяна'): {'completed_attemps': [{'Обучение по продажам': {'accuracy_percent': 89.8,
                                                                                     'correct_answers': 211,
                                                                                     'course_name': 'Обучение '
                                                                                                    'по '
                                                                                                    'продажам',
                                                                                     'date_completed': '2026-05-24T18:51:11.219997',
                                                                                     'lessons_completed': 43,
                                                                                     'total_answers': 235}},
                                                           {'Обучение по продажам': {'accuracy_percent': 46.8,
                                                                                     'correct_answers': 211,
                                                                                     'course_name': 'Обучение '
                                                                                                    'по '
                                                                                                    'продажам',
                                                                                     'date_completed': '2026-06-24T18:51:11.219997',
                                                                                     'lessons_completed': 43,
                                                                                     'total_answers': 235}},
                                                           {'Обучение по продажам': {'accuracy_percent': 38.8,
                                                                                     'correct_answers': 211,
                                                                                     'course_name': 'Обучение '
                                                                                                    'по '
                                                                                                    'продажам',
                                                                                     'date_completed': '2026-05-15T18:51:11.219997',
                                                                                     'lessons_completed': 43,
                                                                                     'total_answers': 235}}],
                                     'not_completed_courses': []},
 ('49728997', 'Кадыйрова Гульнара'): {'completed_attemps': [{'Обучение по продажам': {'accuracy_percent': 88.9,
                                                                                      'correct_answers': 209,
                                                                                      'course_name': 'Обучение '
                                                                                                     'по '
                                                                                                     'продажам',
                                                                                      'date_completed': '2026-05-04T12:50:29.461827',
                                                                                      'lessons_completed': 43,
                                                                                      'total_answers': 235}}],
                                      'not_completed_courses': []},
 ('50076911', 'Павлова Елена'): {'completed_attemps': [],
                                 'not_completed_courses': [{'Обучение по продажам': {'accuracy_percent': 82.9,
                                                                                     'correct_answers': 170,
                                                                                     'lessons_completed': 42,
                                                                                     'total_answers': 205,
                                                                                     'total_lessons': 43}}]},
 ('51490094', 'Шабалин Илья'): {'completed_attemps': [{'Обучение по продажам': {'accuracy_percent': 80.0,
                                                                                'correct_answers': 188,
                                                                                'course_name': 'Обучение '
                                                                                               'по '
                                                                                               'продажам',
                                                                                'date_completed': '2026-04-15T13:14:37.087643',
                                                                                'lessons_completed': 43,
                                                                                'total_answers': 235}}],
                                'not_completed_courses': []},
 ('85179182', 'Канашина Наталья'): {'completed_attemps': [],
                                    'not_completed_courses': [{'Обучение по продажам': {'accuracy_percent': 82.2,
                                                                                        'correct_answers': 37,
                                                                                        'lessons_completed': 7,
                                                                                        'total_answers': 45,
                                                                                        'total_lessons': 43}}]}}
    

